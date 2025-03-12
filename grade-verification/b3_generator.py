import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_b3_file(changes, 
                     template_path="B3 Grade Corrections Template new.xlsx", 
                     output_path="B3_Grade_Corrections_FINAL.xlsx"):
    """
    Writes grade changes to the B3 template starting at row 12.
    Row 11 is used for column headers with your exact requested formatting.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    wb = load_workbook(template_path)
    sheet = wb.active  # or use wb["SheetName"] if your template is named

    # ---------------------------------------------------------------------
    # 1) Write column headers in row 11 with your EXACT formatting
    # ---------------------------------------------------------------------
    # * Fill color: black (#000000)
    # * Text color: #CFB87C
    # * Font: Arial, size 10, not bold
    # * Alignment: Center
    header_map = [
        ("A11", "Student ID"),
        ("B11", "Last Name"),
        ("C11", "First Name"),
        ("D11", "Term"),
        ("E11", "Class Number"),
        ("F11", "Institution"),
        ("G11", "Career"),
        ("H11", "Program"),
        ("I11", "Action"),
        ("J11", "Old Grade"),
        ("K11", "New Grade"),
        ("L11", "Class Subject"),
        ("M11", "Catalog No."),
        ("N11", "Section"),
        ("O11", "Notes"),
    ]

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    # Use Arial size 10, not bold, color #CFB87C
    header_font = Font(name="Arial", size=10, bold=False, color="CFB87C")
    center_align = Alignment(horizontal="center")

    for cell_ref, header_text in header_map:
        cell = sheet[cell_ref]
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # ---------------------------------------------------------------------
    # 2) Start writing data at row 12 (Arial size 10, black text)
    # ---------------------------------------------------------------------
    start_row = 12
    current_row = start_row

    for change in changes:
        sheet[f"A{current_row}"] = change.get('student_id', '')
        sheet[f"B{current_row}"] = change.get('last_name', '')
        sheet[f"C{current_row}"] = change.get('first_name', '')
        sheet[f"D{current_row}"] = change.get('term', '2247')
        sheet[f"E{current_row}"] = change.get('class_number', change.get('crn', ''))
        sheet[f"F{current_row}"] = change.get('institution', 'CUBLD')
        sheet[f"G{current_row}"] = change.get('career', 'GRD3')
        sheet[f"H{current_row}"] = change.get('program', 'MEEM')
        sheet[f"I{current_row}"] = "Grade Correction"
        sheet[f"J{current_row}"] = change.get('old_grade', '')
        sheet[f"K{current_row}"] = change.get('new_grade', '')
        sheet[f"L{current_row}"] = change.get('course_subject', '')
        sheet[f"M{current_row}"] = change.get('course_number', '')
        sheet[f"N{current_row}"] = change.get('section', '')
        sheet[f"O{current_row}"] = change.get('notes', '')

        # Data rows: Arial size 10, black text (#000000), center aligned
        data_font = Font(name="Arial", size=10, bold=False, color="000000")
        for col in range(1, 16):  # A=1 through O=15
            cell = sheet.cell(row=current_row, column=col)
            cell.font = data_font
            cell.alignment = center_align

        current_row += 1

    wb.save(output_path)
    print(f"B3 file generated: {output_path}")
